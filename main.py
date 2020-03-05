import discord
from discord.ext import commands
from discord.utils import get
import random as rnd
import openpyxl
import asyncio
import datetime as dt


with open("token.txt", 'r') as f:
    token = f.readline()

with open("ids.txt", 'r') as f:
    roleMessageID = int(f.readline()[:-1])
    roleMessageResetID = int(f.readline())

client = discord.Client()
bot = commands.Bot(command_prefix='!', case_insensitive = True)
bot.remove_command('help')

wb = openpyxl.load_workbook('2020 Member List.xlsx')
ws = wb['Sheet1']


@bot.event
async def on_ready():
    global server, generalChannel, ruleChannel, botChannel, botMember, execBotChannel, execRole
    server = get(bot.guilds, name='UoATGC')
    generalChannel = get(server.channels, name='general')
    ruleChannel = get(server.channels, name='rules')
    botChannel = get(server.channels, name='bot-commands')
    botMember = get(server.members, id=bot.user.id)
    execBotChannel = get(server.channels, name='exec-bot')
    execRole = get(server.roles, name='Exec')
    print('Bot is ready')
    msg = await ruleChannel.fetch_message(roleMessageID)
    for reaction in msg.reactions:
        if reaction.emoji == '🎲':
            role = get(server.roles, name='Board Games')
            for user in await reaction.users().flatten():
                if user is not botMember: await user.add_roles(role)
        if reaction.emoji == '⚔️':
            role = get(server.roles, name='RPGs')
            for user in await reaction.users().flatten():
                if user is not botMember: await user.add_roles(role)
        if reaction.emoji == '🃏':
            role = get(server.roles, name='TCGs')
            for user in await reaction.users().flatten():
                if user is not botMember: await user.add_roles(role)
    msg = await ruleChannel.fetch_message(roleMessageResetID)
    for reaction in msg.reactions:
        if reaction.emoji == '🎲':
            role = get(server.roles, name='Board Games')
            for user in await reaction.users().flatten():
                if user is not botMember: await user.add_roles(role)
        if reaction.emoji == '⚔️':
            role = get(server.roles, name='RPGs')
            for user in await reaction.users().flatten():
                if user is not botMember: await user.add_roles(role)
        if reaction.emoji == '🃏':
            role = get(server.roles, name='TCGs')
            for user in await reaction.users().flatten():
                if user is not botMember: await user.add_roles(role)


@bot.command()
async def colours(ctx):
    if ctx.channel is not botChannel: return
    msg = await botChannel.send('React for a pastel coloured role:')
    reactions = ['❤️','🧡','💛','💚','💙','💜','🖤']
    for reaction in reactions:
        await msg.add_reaction(reaction)
    msg = await botChannel.send('React for a light coloured role:')
    reactions = ['🔴','🟠','🟡','🟢','🔵','🟣','⚫']
    for reaction in reactions:
        await msg.add_reaction(reaction)
    msg = await botChannel.send('React for a dark coloured role:')
    reactions = ['🟥','🟧','🟨','🟩','🟦','🟪','⬛']
    for reaction in reactions:
        await msg.add_reaction(reaction)

welcomeMsg = '''Welcome to the TGC Discord! Be sure to read the server rules
Reply with your Uni ID to be given the member role, or react with 🎲
if you aren't from UoA to alert one of the exec members to message you
'''
@bot.event
async def on_member_join(user):
    msg = await user.send(welcomeMsg)
    await execBotChannel.send(user.name + ' has joined the server')
    await msg.add_reaction('🎲')


@bot.event
async def on_message(msg):
    await bot.process_commands(msg)
    if msg.guild is None:
        member = get(server.members, id=msg.author.id)
        if msg.author is bot.user or member is None: return
        elif get(member.roles, name='Member') is not None: return
        i = 2
        try:
            while ws.cell(i, 1).value is not None:
                if msg.content.strip() == str(ws.cell(i, 2).value):
                    memberRole = get(server.roles, name='Member')
                    await member.add_roles(memberRole)
                    await generalChannel.send('Welcome, ' + member.mention + '!')
                    await execBotChannel.send(member.display_name + '/' + ws.cell(i,1).value + ' has joined as a member')
                    return
                i += 1
            await member.send('ID not found, react to ask exec to let you in manually')
        except ValueError:
            await member.send('ID not found, react to ask exec to let you in manually')
        return
    if botMember.mentioned_in(msg):
        await msg.channel.send('Bruh that\'s cringe')


@bot.event
async def on_raw_reaction_add(payload):
    guild = get(bot.guilds, id=payload.guild_id)
    emoji = payload.emoji.name
    if payload.user_id == botMember.id: return
    if guild is None:
        member = get(server.members, id=payload.user_id)
        if member is None: return
        if get(member.roles, name='Member') is None and emoji == '🎲':
            await execBotChannel.send(execRole.mention + ' ' + member.display_name + ' wants to be manually made a member of the server')
        return
    channel = get(server.channels, id=payload.channel_id)
    member = payload.member
    bgRole = get(server.roles, name='Board Games')
    rpgRole = get(server.roles, name='RPGs')
    tcgRole = get(server.roles, name='TCGs')

    if get(member.roles, name='Member') is None: return
    if payload.message_id == roleMessageID:
        if emoji == '🎲':
            await member.add_roles(bgRole)
        elif emoji == '⚔️':
            await member.add_roles(rpgRole)
        elif emoji == '🃏':
            await member.add_roles(tcgRole)
        msg = await ruleChannel.fetch_message(roleMessageResetID)
        users = None
        for reaction in msg.reactions:
            if reaction.emoji == emoji:
                users = await reaction.users().flatten()
                break
        if users is not None:
            if member in users:
                await reaction.remove(member)
    if payload.message_id == roleMessageResetID:
        if emoji == '🎲':
            await member.add_roles(bgRole)
        elif emoji == '⚔️':
            await member.add_roles(rpgRole)
        elif emoji == '🃏':
            await member.add_roles(tcgRole)
        msg = await ruleChannel.fetch_message(roleMessageID)
        users = None
        for reaction in msg.reactions:
            if reaction.emoji == emoji:
                users = await reaction.users().flatten()
                break
        if users is not None:
            if member in users:
                await reaction.remove(member)
    if channel is botChannel:
        msg = await botChannel.fetch_message(payload.message_id)
        if msg.content.startswith('React for a ') and msg.content.endswith(' coloured role:'):
            if 'colour_' in [role.name[:7] for role in member.roles]: return
            reactions = ['❤️','🧡','💛','💚','💙','💜','🖤',
            '🔴','🟠','🟡','🟢','🔵','🟣','⚫',
            '🟥','🟧','🟨','🟩','🟦','🟪','⬛']
            names = ['colour_red1','colour_orange1','colour_yellow1','colour_green1','colour_blue1','colour_purple1','colour_grey1',
            'colour_red2','colour_orange2','colour_yellow2','colour_green2','colour_blue2','colour_purple2','colour_grey2',
            'colour_red3','colour_orange3','colour_yellow3','colour_green3','colour_blue3','colour_purple3','colour_grey3']
            hexes = [0xffa0a0,0xffbf93,0xffeaa9,0xa6f1c8,0xa2defd,0xc09df0,0xe4cde2,
            0xce3e3e,0xce7235,0xd4ab26,0x3dc482,0x398fd1,0x8545db,0x928b92,
            0x8a1313,0x974a17,0x917311,0x0e6d34,0x103f8d,0x621997,0x464546]
            triples = [[reactions[i],names[i],hexes[i]] for i in range(len(reactions))]
            for triple in triples:
                if emoji == triple[0]:
                    role = get(server.roles,name=triple[1])
                    if role is None:
                        role = await server.create_role(name=triple[1], colour=discord.Colour(triple[2]))
                    await member.add_roles(role)
                    return
    if emoji == '🍆' or emoji == '🍑':
        msg = await channel.fetch_message(payload.message_id)
        if msg.author is botMember:
            await member.send('😉')

@bot.event
async def on_member_update(before, after):
    if after.roles != before.roles:
        for role in after.roles:
            if role in before.roles: continue
            if role.name.startswith('colour_') and role.position == 1:
                try:
                    position = get(server.roles,name='Member').position
                except NameError: return
                for i in range(2,position+1):
                    await server.roles[i].edit(position=i-1)
                await role.edit(position=position)
                return


@bot.event
async def on_raw_reaction_remove(payload):
    guild = get(bot.guilds, id=payload.guild_id)
    if payload.user_id == botMember.id or guild is None: return
    emoji = payload.emoji.name
    channel = get(server.channels, id=payload.channel_id)
    member = get(server.members, id=payload.user_id)
    bgRole = get(server.roles, name='Board Games')
    rpgRole = get(server.roles, name='RPGs')
    tcgRole = get(server.roles, name='TCGs')

    if payload.message_id == roleMessageID:
        msg = await ruleChannel.fetch_message(roleMessageResetID)
        for reaction in msg.reactions:
            if reaction.emoji == emoji:
                users = await reaction.users().flatten()
                break
        if member in users: return
        if emoji == '🎲':
            await member.remove_roles(bgRole)
        elif emoji == '⚔️':
            await member.remove_roles(rpgRole)
        elif emoji == '🃏':
            await member.remove_roles(tcgRole)
    if payload.message_id == roleMessageResetID:
        msg = await ruleChannel.fetch_message(roleMessageID)
        for reaction in msg.reactions:
            if reaction.emoji == emoji:
                users = await reaction.users().flatten()
                break
        if member in users: return
        if emoji == '🎲':
            await member.remove_roles(bgRole)
        elif emoji == '⚔️':
            await member.remove_roles(rpgRole)
        elif emoji == '🃏':
            await member.remove_roles(tcgRole)
    if channel is botChannel:
        msg = await channel.fetch_message(payload.message_id)
        if msg.content.startswith('React for a ') and msg.content.endswith(' coloured role:'):
            reactions = ['❤️','🧡','💛','💚','💙','💜','🖤',
            '🔴','🟠','🟡','🟢','🔵','🟣','⚫',
            '🟥','🟧','🟨','🟩','🟦','🟪','⬛']
            names = ['colour_red1','colour_orange1','colour_yellow1','colour_green1','colour_blue1','colour_purple1','colour_grey1',
            'colour_red2','colour_orange2','colour_yellow2','colour_green2','colour_blue2','colour_purple2','colour_grey2',
            'colour_red3','colour_orange3','colour_yellow3','colour_green3','colour_blue3','colour_purple3','colour_grey3']
            pairs = [[reactions[i],names[i]] for i in range(len(reactions))]
            for pair in pairs:
                if emoji == pair[0]:
                    role = get(server.roles,name=pair[1])
                    if len(role.members) == 1: await role.delete()
                    else: await member.remove_roles(role)


async def reset():
    while True:
        now = dt.datetime.now()
        midnight = dt.datetime.combine(dt.datetime.today() + dt.timedelta(days=1), dt.time.min)
        await asyncio.sleep((midnight-now).total_seconds() + 60)

        if rnd.random() < 0.01:
            await botMember.edit(nick='Carlos Webster, UoATGC President')
        else:
            await botMember.edit(nick=None)

        if dt.datetime.today().weekday() != 6: continue
        msg = await ruleChannel.fetch_message(roleMessageResetID)
        bgRole = get(server.roles, name='Board Games')
        rpgRole = get(server.roles, name='RPGs')
        tcgRole = get(server.roles, name='TCGs')
        for reaction in msg.reactions:
            emoji = reaction.emoji 
            for member in await reaction.users().flatten():
                if member.id == botMember.id: continue
                if emoji == '🎲':
                    await member.remove_roles(bgRole)
                elif emoji == '⚔️':
                    await member.remove_roles(rpgRole)
                elif emoji == '🃏':
                    await member.remove_roles(tcgRole)
                await reaction.remove(member)


bot.loop.create_task(reset())
bot.run(token)